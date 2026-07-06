import openpyxl

# Real emails from database, keyed by student ID (SL. NO.)
email_map = {
    1: "abhishek.singh@gmail.com",
    2: "abhishek.kumar@gmail.com",
    3: "abu.luqmani@gmail.com",
    4: "sahil.kumar@gmail.com",
    5: "biswarup.chatterjee@gmail.com",
    6: "aman.kumar@gmail.com",
    7: "nandini.singh@gmail.com",
    8: "monami.rana@gmail.com",
    9: "mojammil.ansari@gmail.com",
    10: "labanya.saha@gmail.com",
    11: "kushal.biswas@gmail.com",
    12: "kumar.mridul@gmail.com",
    13: "koyena.dutta@gmail.com",
    14: "joydeep.das@gmail.com",
    15: "dipanwita.pal@gmail.com",
    16: "dipannita.mukherjee@gmail.com",
    18: "chandika.sarkar@gmail.com",
    19: "priyanshu.virendra@gmail.com",
    20: "ayana.baidya@gmail.com",
    21: "asmita.das@gmail.com",
    22: "arpan.pan@gmail.com",
    23: "arnab.gupta@gmail.com",
    24: "arkadeep.pathak@gmail.com",
    25: "arijit.ghosh@gmail.com",
    26: "rishu.modi@gmail.com",
    27: "anushka.singh@gmail.com",
    28: "aditya.raj@gmail.com",
    29: "ankit.gupta@gmail.com",
    30: "anangsha.mitra@gmail.com",
    31: "amrit.pramanik@gmail.com",
    32: "dayyanul.haque@gmail.com",
    33: "soumyajit.jana@gmail.com",
    34: "pukar.sharma@gmail.com",
    35: "karunya.raj@gmail.com",
    36: "abhishek.saha@gmail.com",
    37: "suprakash.roy@gmail.com",
    38: "sumit.dubey@gmail.com",
    39: "suhani.kundu@gmail.com",
    40: "srijan.paul@gmail.com",
    41: "srijan.das@gmail.com",
    42: "srijan.bhattacharyya@gmail.com",
    43: "souvik.pachal@gmail.com",
    44: "souvik.dutta@gmail.com",
    45: "nilesh.choudhury@gmail.com",
    46: "sourasish.chatterjee@gmail.com",
    47: "anurag.chowdhury@gmail.com",
    48: "soumen.gorai@gmail.com",
    49: "soujanya.khan@gmail.com",
    50: "sohan.samanta@gmail.com",
    51: "shruti.kumari@gmail.com",
    52: "sayan.manna@gmail.com",
    53: "sanchari.roy@gmail.com",
    54: "saheli.majumder@gmail.com",
    55: "rishi.burnwal@gmail.com",
    56: "rhitwika.poddar@gmail.com",
    57: "ravi.kumar@gmail.com",
    58: "rahul.ghosh@gmail.com",
    59: "ragini.shaw@gmail.com",
    60: "sourav.ojha@gmail.com",
    61: "anwesa.maji@gmail.com",
    62: "aditya.verma@gmail.com",
    63: "aditya.singh@gmail.com",
    64: "arin.karmakar@gmail.com",
    65: "chandan.raj@gmail.com",
    66: "rohan.mandal@gmail.com",
    67: "saikat.das@gmail.com",
    68: "sumit.kumar@gmail.com",
    69: "aditya.sharma@gmail.com",
    70: "priya.verma@gmail.com",
    71: "rohan.gupta@gmail.com",
    72: "neha.singh@gmail.com",
    73: "arjun.patel@gmail.com",
    74: "divya.nair@gmail.com",
    75: "vikram.kumar@gmail.com",
    76: "sneha.iyer@gmail.com",
}

xlsx_path = r"d:\EduConnect\FYP 12\dataset\Student_DataSet.xlsx"
out_path  = r"d:\EduConnect\FYP 12\dataset\Student_DataSet_updated.xlsx"

wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

# Find column indices from header row (row 1)
header = {cell.value: cell.column for cell in ws[1] if cell.value}
print("Headers found:", list(header.keys()))

sl_col = None
email_col = None
for h, col in header.items():
    h_str = str(h).strip().upper()
    if 'SL' in h_str or 'S.NO' in h_str or 'S NO' in h_str:
        sl_col = col
    if 'EMAIL' in h_str or 'GMAIL' in h_str:
        email_col = col

print(f"SL.NO column: {sl_col}, EMAIL column: {email_col}")

if not sl_col or not email_col:
    print("ERROR: Could not find required columns")
    exit(1)

updated = 0
for row in ws.iter_rows(min_row=2):
    sl_cell = row[sl_col - 1]
    email_cell = row[email_col - 1]
    
    if sl_cell.value is None:
        continue
    
    try:
        sl_no = int(sl_cell.value)
    except (ValueError, TypeError):
        continue
    
    if sl_no in email_map:
        old_email = email_cell.value
        email_cell.value = email_map[sl_no]
        print(f"  Row {sl_cell.row}: ID={sl_no} → {old_email} → {email_map[sl_no]}")
        updated += 1

wb.save(out_path)
print(f"\nDone! Updated {updated} email addresses.")
print(f"Saved to: {out_path}")
print("Please close Student_DataSet.xlsx in WPS and replace it with the updated file.")
